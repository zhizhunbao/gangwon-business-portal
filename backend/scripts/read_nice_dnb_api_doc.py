"""
读取 Nice D&B API 文档 Excel 文件
"""
import sys
from pathlib import Path
import openpyxl

# 添加项目路径
backend_dir = Path(__file__).parent.parent
project_root = backend_dir.parent
sys.path.insert(0, str(backend_dir))

# Excel 文件路径
excel_path = project_root / "docs" / "(NICEDNB)API_251210.xlsx"

def read_excel_file():
    """读取 Excel 文件并显示内容"""
    if not excel_path.exists():
        print(f"❌ 文件不存在: {excel_path}")
        return
    
    print(f"📖 正在读取: {excel_path}")
    print("=" * 80)
    
    try:
        # 打开工作簿
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        
        # 显示所有工作表名称
        print(f"\n📋 工作表列表 ({len(workbook.sheetnames)} 个):")
        for idx, sheet_name in enumerate(workbook.sheetnames, 1):
            print(f"  {idx}. {sheet_name}")
        
        # 读取每个工作表
        for sheet_name in workbook.sheetnames:
            print("\n" + "=" * 80)
            print(f"📄 工作表: {sheet_name}")
            print("=" * 80)
            
            sheet = workbook[sheet_name]
            
            # 显示工作表信息
            print(f"  行数: {sheet.max_row}")
            print(f"  列数: {sheet.max_column}")
            
            # 读取前几行数据（最多显示前 20 行）
            max_rows_to_show = 20
            print(f"\n  前 {min(max_rows_to_show, sheet.max_row)} 行数据:\n")
            
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if row_idx > max_rows_to_show:
                    print(f"  ... (还有 {sheet.max_row - max_rows_to_show} 行)")
                    break
                
                # 过滤空行
                if any(cell is not None for cell in row):
                    # 格式化输出
                    row_data = []
                    for cell in row:
                        if cell is None:
                            row_data.append("")
                        else:
                            # 截断过长的单元格内容
                            cell_str = str(cell)
                            if len(cell_str) > 50:
                                cell_str = cell_str[:47] + "..."
                            row_data.append(cell_str)
                    
                    print(f"  行 {row_idx:3d}: {' | '.join(row_data)}")
        
        workbook.close()
        print("\n" + "=" * 80)
        print("✅ 读取完成")
        
    except Exception as e:
        print(f"❌ 读取文件时出错: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    read_excel_file()

