"""
Export service.

Provides functionality to export data to Excel and CSV formats.
"""
import io
from typing import Any, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import csv

from ..logger import get_logger

logger = get_logger(__name__)


class ExportService:
    """Service for exporting data to Excel and CSV formats."""

    @staticmethod
    def export_to_excel(
        data: list[dict[str, Any]],
        sheet_name: str = "Data",
        headers: Optional[list[str]] = None,
        title: Optional[str] = None,
    ) -> bytes:
        """
        Export data to Excel format.

        Args:
            data: List of dictionaries containing data to export
            sheet_name: Name of the Excel sheet
            headers: Optional list of header names (if None, uses dict keys from first row)
            title: Optional title row

        Returns:
            Excel file as bytes
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Set default headers if not provided
            if not headers and data:
                headers = list(data[0].keys())

            if not headers:
                # Empty data, return empty workbook
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                return output.getvalue()

            # Add title row if provided
            row_num = 1
            if title:
                ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
                title_cell = ws["A1"]
                title_cell.value = title
                title_cell.font = Font(size=14, bold=True)
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                title_cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                title_cell.font = Font(size=14, bold=True, color="FFFFFF")
                row_num = 2

            # Add headers
            header_fill = PatternFill(
                start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
            )
            header_font = Font(bold=True, size=11)

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Add data rows
            data_row = row_num + 1
            for row_data in data:
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=data_row, column=col_num)
                    value = row_data.get(header, "")
                    
                    # Format datetime objects
                    if isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    elif value is None:
                        value = ""
                    elif isinstance(value, (list, dict)):
                        value = str(value)
                    
                    cell.value = value
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

                data_row += 1

            # Auto-adjust column widths
            for col_num, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_num)
                max_length = len(str(header))
                
                # Check data rows for max length
                for row in ws.iter_rows(min_row=data_row - len(data), max_row=data_row - 1, min_col=col_num, max_col=col_num):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                
                # Set column width (with some padding)
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            logger.info(
                f"Exported {len(data)} rows to Excel",
                extra={
                    "module_name": __name__,
                    "sheet_name": sheet_name,
                    "row_count": len(data),
                },
            )
            
            return output.getvalue()

        except Exception as e:
            logger.error(
                f"Error exporting to Excel: {str(e)}",
                exc_info=True,
                extra={"export_module": __name__},
            )
            raise

    @staticmethod
    def export_to_csv(
        data: list[dict[str, Any]],
        headers: Optional[list[str]] = None,
    ) -> str:
        """
        Export data to CSV format.

        Args:
            data: List of dictionaries containing data to export
            headers: Optional list of header names (if None, uses dict keys from first row)

        Returns:
            CSV content as string with UTF-8 BOM for Excel compatibility
        """
        try:
            if not data:
                return ""

            # Set default headers if not provided
            if not headers:
                headers = list(data[0].keys())

            # Create CSV in memory
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
            writer.writeheader()

            for row_data in data:
                # Format datetime objects
                formatted_row = {}
                for key, value in row_data.items():
                    if isinstance(value, datetime):
                        formatted_row[key] = value.strftime("%Y-%m-%d %H:%M:%S")
                    elif value is None:
                        formatted_row[key] = ""
                    else:
                        formatted_row[key] = value
                writer.writerow(formatted_row)

            csv_content = output.getvalue()
            output.close()

            logger.info(
                f"Exported {len(data)} rows to CSV",
                extra={
                    "module_name": __name__,
                    "row_count": len(data),
                },
            )

            # Add UTF-8 BOM for Excel compatibility with Korean/Chinese characters
            return "\ufeff" + csv_content

        except Exception as e:
            logger.error(
                f"Error exporting to CSV: {str(e)}",
                exc_info=True,
                extra={"export_module": __name__},
            )
            raise

